import os
import numpy as np
import pandas as pd
from scipy.signal import find_peaks, savgol_filter
from data_parser import parse_wafer_data
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==========================================================
# [설정] 아웃라이어(Outlier) 및 에러 판단 기준
# ==========================================================
THRESHOLD = {
    'IL_MIN': -20.0,  # IL이 -20dB보다 낮으면 에러 (너무 큰 손실)
    'ER_MIN': 10.0,  # ER이 10dB보다 낮으면 아웃라이어
    'VPIL_MIN': 0.2,  # VpiL이 0.2Vcm 미만이면 에러
    'VPIL_MAX': 3.0  # VpiL이 3.0Vcm 초과하면 에러
}

zip_path = "../dat/HY202103.zip"
save_path = "../res/Process_result.xlsx"
os.makedirs("../res", exist_ok=True)

target_wafers = ['D07', 'D08', 'D23', 'D24']
L_length = 0.05  # phase shifter length [cm]


def q_sub(x, y):
    """valley 주변 3점으로 subpixel 위치 정밀 보정"""
    if len(x) < 3: return x[np.argmin(y)]
    idx = np.argmin(y)
    if idx == 0 or idx == len(x) - 1: return x[idx]
    c = np.polyfit(x[idx - 1:idx + 2], y[idx - 1:idx + 2], 2)
    return -c[1] / (2 * c[0]) if abs(c[0]) > 1e-12 else x[idx]


def check_status(row):
    """데이터의 오류나 아웃라이어를 체크하고 이유를 반환"""
    reasons = []
    if pd.isna(row['IL_dB']):
        reasons.append("IL 데이터 없음")
    elif row['IL_dB'] < THRESHOLD['IL_MIN']:
        reasons.append(f"IL 손실 과다(<{THRESHOLD['IL_MIN']})")

    if pd.isna(row['ER_dB']):
        reasons.append("ER 데이터 없음")
    elif row['ER_dB'] < THRESHOLD['ER_MIN']:
        reasons.append(f"ER 낮음(<{THRESHOLD['ER_MIN']})")

    if pd.isna(row['VpiL_Vcm']):
        reasons.append("VpiL 계산 실패")
    elif row['VpiL_Vcm'] < THRESHOLD['VPIL_MIN'] or row['VpiL_Vcm'] > THRESHOLD['VPIL_MAX']:
        reasons.append(f"VpiL 범위 초과({THRESHOLD['VPIL_MIN']}~{THRESHOLD['VPIL_MAX']})")

    if not reasons:
        return "정상", ""
    else:
        return "이상 발생", ", ".join(reasons)


print("🚀 데이터 분석 및 엑셀 리포트 생성을 시작합니다...")
summary_list = []
count = 0

# --- 데이터 추출 및 계산 루프 ---
for d in parse_wafer_data(zip_path, target_wafers):
    wafer, band, c, r = d['wafer_id'], d['band'], d['die_c'], d['die_r']
    radius = np.sqrt(c ** 2 + r ** 2)

    # [중요] 경로 매칭을 위해 날짜 정보 가져오기
    date_str = d.get('date', 'Unknown_Date')

    # --- 0. Baseline (Ref) Fitting ---
    m = (d['ref_data']['wl'] >= d['wl_min']) & (d['ref_data']['wl'] <= d['wl_max'])
    v_ref_wl, v_ref_il = d['ref_data']['wl'][m], d['ref_data']['il'][m]
    if len(v_ref_wl) < 31: continue

    poly_func = np.poly1d(np.polyfit(v_ref_wl, savgol_filter(v_ref_il, 31, 3), 3))

    # --- 1. IL Calculation ---
    max_peak = float('-inf')
    for b in d['bias_data_list']:
        if b['bias'] is None: continue
        mb = (b['wl'] >= d['wl_min']) & (b['wl'] <= d['wl_max'])
        if np.any(mb):
            max_peak = max(max_peak, np.max(b['il'][mb]))
    il_val = max_peak if max_peak != float('-inf') else np.nan

    # --- 2. ER Calculation ---
    max_er = 0.0
    for b in d['bias_data_list']:
        if b['bias'] is None: continue
        mb = (b['wl'] >= d['wl_min']) & (b['wl'] <= d['wl_max'])
        v_wl, v_il = b['wl'][mb], b['il'][mb]
        if len(v_wl) == 0: continue

        flat_il = v_il - poly_func(v_wl)
        peaks, _ = find_peaks(flat_il, prominence=3.0, distance=30)

        if len(peaks) >= 2:
            trend = np.poly1d(np.polyfit(v_wl[peaks], flat_il[peaks], 1))
            final_il = flat_il - trend(v_wl)
        else:
            final_il = flat_il

        er = np.percentile(final_il, 99) - np.percentile(final_il, 1)
        max_er = max(max_er, er)
    er_val = max_er if max_er > 0 else np.nan

    # --- 3. VpiL Calculation ---
    vpil_val = np.nan
    z_data = next((b for b in d['bias_data_list'] if b['bias'] == 0.0), None)

    if z_data:
        m0 = (z_data['wl'] >= d['wl_min']) & (z_data['wl'] <= d['wl_max'])
        w0, i0 = z_data['wl'][m0], z_data['il'][m0]
        flat0 = savgol_filter(i0, 31, 3) - poly_func(w0)
        v0, _ = find_peaks(-flat0, prominence=0.3, distance=20)

        if len(v0) >= 2:
            fsr = np.mean(np.diff(w0[v0]))
            cwl = w0[v0[np.argmin(np.abs(w0[v0] - d['target_wl']))]]

            volts, p_pi = [], []
            sh = fsr / 2.5
            for b in d['bias_data_list']:
                if b['bias'] is None: continue
                w, i = b['wl'], b['il']
                mb = (w >= d['wl_min']) & (w <= d['wl_max'])
                wb, ib = w[mb], i[mb]

                mloc = (wb >= cwl - sh) & (wb <= cwl + sh)
                if np.sum(mloc) < 5: continue

                flat = savgol_filter(ib[mloc], 11, 3) - poly_func(wb[mloc])
                vwl = q_sub(wb[mloc], flat)
                volts.append(b['bias'])
                p_pi.append(2.0 * (vwl - cwl) / fsr)

            if len(volts) >= 5:
                volts, p_pi = np.array(volts), np.array(p_pi)

                fit = np.poly1d(np.polyfit(volts, p_pi, 2))
                derivative = np.polyder(fit)

                slope_at_0 = np.abs(derivative(0.0))
                slope_at_0 = max(slope_at_0, 1e-5)
                vpil_0V = L_length / slope_at_0

                if 0.1 <= vpil_0V <= 10.0:
                    vpil_val = vpil_0V

    # 리스트에 계산된 값 추가 (Date 항목 추가)
    summary_list.append({
        'Wafer': wafer, 'Band': band, 'Date': date_str, 'Column': c, 'Row': r, 'Radius': radius,
        'IL_dB': il_val, 'ER_dB': er_val, 'VpiL_Vcm': vpil_val
    })

    count += 1
    if count % 200 == 0:
        print(f"현재 {count}개 Die 분석 완료...")

# --- DataFrame 변환 및 상태 체크 ---
df = pd.DataFrame(summary_list)

# 중복 데이터 최소화 처리 (Date 항목도 보존하기 위해 groupby 대상에 포함)
df = df.groupby(['Wafer', 'Band', 'Date', 'Column', 'Row', 'Radius'], as_index=False).min(numeric_only=True)

# 정상/이상 여부 판단
df['Status'], df['Reason'] = zip(*df.apply(check_status, axis=1))

# 링크를 적용할 'Folder_Link' 가상 컬럼 추가 (엑셀에서 수식으로 변환 예정)
df['Folder_Link'] = "Click"

print("엑셀 파일을 저장하고 하이퍼링크 서식을 적용하는 중...")

# --- 엑셀 저장 및 서식 적용 ---
with pd.ExcelWriter(save_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Analysis_Result')
    worksheet = writer.sheets['Analysis_Result']

    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    # 하이퍼링크용 폰트 스타일 (파란색 + 밑줄)
    link_font = Font(color='0563C1', underline='single')

    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # 각 열의 인덱스 매핑 확인
    col_idx = {col: i for i, col in enumerate(df.columns, 1)}

    for col_num, column_title in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

        for row_num in range(2, len(df) + 2):
            data_cell = worksheet.cell(row=row_num, column=col_num)
            data_cell.border = border

            # 1. 하이퍼링크 수식 적용 (Folder_Link 열인 경우)
            if column_title == 'Folder_Link':
                # 같은 행의 다른 셀에서 파일 경로에 필요한 변수 추출
                w_id = worksheet.cell(row=row_num, column=col_idx['Wafer']).value
                d_str = worksheet.cell(row=row_num, column=col_idx['Date']).value
                die_c = worksheet.cell(row=row_num, column=col_idx['Column']).value
                die_r = worksheet.cell(row=row_num, column=col_idx['Row']).value

                # 상대 경로 생성 (엑셀 파일 위치 기준: ../res/Wafer/날짜/좌표)
                # 윈도우 탐색기에서 바로 열리도록 역슬래시(\) 조합으로 세팅
                target_folder = f".\\{w_id}\\{d_str}\\C{die_c}_R{die_r}"

                # HYPERLINK 수식 삽입
                data_cell.value = f'=HYPERLINK("{target_folder}", "📂 바로가기")'
                data_cell.font = link_font
                data_cell.alignment = Alignment(horizontal='center')

            else:
                # 일반 숫자 포맷 고정
                if column_title in ['IL_dB', 'ER_dB', 'VpiL_Vcm']:
                    data_cell.number_format = '0.000'

            # 2. 에러 행 하이라이트 (링크 열은 파란색 유지를 위해 제외)
            status_value = worksheet.cell(row=row_num, column=col_idx['Status']).value
            if status_value == "이상 발생" and column_title != 'Folder_Link':
                data_cell.fill = error_fill

    # 스마트 열 너비 조절
    for col in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                text = str(cell.value)
                adjusted_length = sum(2 if ord(c) > 127 else 1 for c in text)
                max_length = max(max_length, adjusted_length)
        worksheet.column_dimensions[column_letter].width = max_length + 3

print(f"✅ 하이퍼링크 리포트 완료: {save_path}")